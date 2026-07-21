"""Parse a screener.in "Export to Excel" workbook into tidy line items.

The export has a single sheet ("Data Sheet") laid out as stacked sections:

    PROFIT & LOSS
              Mar-2022   Mar-2023   Mar-2024     <- period header row
    Sales      792756     976524    1000122
    ...
    QUARTERS
              Dec-2023   Mar-2024   Jun-2024
    Sales      ...
    BALANCE SHEET
    ...
    CASH FLOW:
    ...

We locate each section header, read the period row beneath it, then emit one
record per (line_item, period) cell. Row labels are kept verbatim so we stay
robust to bank-vs-manufacturer differences (see models.py design note).
"""
from __future__ import annotations

import calendar
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

# Section header (normalised) -> (statement key, period_type)
_SECTIONS: dict[str, tuple[str, str]] = {
    "PROFIT & LOSS": ("pnl", "annual"),
    "QUARTERS": ("quarters", "quarter"),
    "BALANCE SHEET": ("balance_sheet", "annual"),
    "CASH FLOW": ("cash_flow", "annual"),
}
# Sections we recognise as boundaries but do not ingest (yet).
_IGNORED_SECTIONS = {"DERIVED", "PRICE"}

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


@dataclass
class LineRecord:
    statement: str
    period_type: str
    period_end: date
    line_item: str
    value: float | None


@dataclass
class ParsedCompany:
    lines: list[LineRecord] = field(default_factory=list)

    def by_statement(self, statement: str) -> list[LineRecord]:
        return [r for r in self.lines if r.statement == statement]


def _norm_header(value) -> str | None:
    if not isinstance(value, str):
        return None
    return value.strip().rstrip(":").upper() or None


def _parse_period(value) -> date | None:
    """Coerce a period-header cell to the last day of its month."""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        last = calendar.monthrange(value.year, value.month)[1]
        return date(value.year, value.month, last)
    if isinstance(value, str):
        text = value.strip().lower().replace("'", "-").replace(" ", "-")
        # Accept "Mar-2024", "Mar-24", "mar 2024".
        parts = [p for p in text.replace("/", "-").split("-") if p]
        if len(parts) >= 2 and parts[0][:3] in _MONTHS:
            month = _MONTHS[parts[0][:3]]
            year_txt = parts[1]
            if year_txt.isdigit():
                year = int(year_txt)
                if year < 100:
                    year += 2000
                last = calendar.monthrange(year, month)[1]
                return date(year, month, last)
    return None


def _parse_value(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("%", "").replace("₹", "")
        if cleaned in {"", "-", "NA", "N/A"}:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _load_rows(path: Path) -> list[tuple]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Data Sheet"] if "Data Sheet" in wb.sheetnames else wb.active
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def parse_workbook(path: Path) -> ParsedCompany:
    rows = _load_rows(path)
    parsed = ParsedCompany()

    i = 0
    n = len(rows)
    while i < n:
        row = rows[i]
        header = _norm_header(row[0]) if row else None

        if header in _IGNORED_SECTIONS:
            i += 1
            continue
        if header not in _SECTIONS:
            i += 1
            continue

        statement, period_type = _SECTIONS[header]

        # The next row holds the period headers.
        if i + 1 >= n:
            break
        period_row = rows[i + 1]
        periods: list[tuple[int, date]] = []
        for col, cell in enumerate(period_row):
            if col == 0:
                continue
            period = _parse_period(cell)
            if period is not None:
                periods.append((col, period))

        # Data rows run until the next recognised section or a blank label.
        j = i + 2
        seen_labels: set[str] = set()
        while j < n:
            data_row = rows[j]
            label_raw = data_row[0] if data_row else None
            norm = _norm_header(label_raw)
            if norm in _SECTIONS or norm in _IGNORED_SECTIONS:
                break
            if not isinstance(label_raw, str) or not label_raw.strip():
                j += 1
                # A single blank row inside a section is a separator, but two
                # in a row (plus no upcoming label) usually means section end.
                continue

            label = label_raw.strip()
            # Disambiguate repeated labels (e.g. "Total" in the balance sheet).
            if label in seen_labels:
                label = f"{label} (2)"
            seen_labels.add(label)

            for col, period in periods:
                value = _parse_value(data_row[col] if col < len(data_row) else None)
                parsed.lines.append(
                    LineRecord(statement, period_type, period, label, value)
                )
            j += 1

        i = j

    return parsed
